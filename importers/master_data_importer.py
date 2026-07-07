from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from database import get_db_connection
from logging_config import get_logger


logger = get_logger(__name__)

PRODUCT_SHEET = "颜色尺码(&X)"
MASTER_SHEET = "客户设置"
PRODUCT_HEADER_ROW = 3
MASTER_HEADER_ROW = 3

DESIRED_TABLE_COLUMNS = {
    "dim_product": [
        "product_code",
        "product_name",
        "color_detail",
        "size_detail",
        "unit_name",
        "year_code",
        "major_category_code",
        "major_category_name",
        "season_code",
        "season_name",
        "brand_code",
        "brand_name",
        "supplier_code",
        "supplier_name",
        "category_code",
        "category_name",
        "production_factory_code",
        "production_factory_name",
        "designer_code",
        "designer_name",
        "launch_wave_code",
        "launch_wave_name",
        "main_material_code",
        "main_material_name",
        "series_code",
        "series_name",
        "standard_purchase_price",
        "purchase_price_1",
        "purchase_price_2",
        "cost_price",
        "standard_retail_price",
        "foreign_trade_price",
        "shipping_price",
        "retail_price_3",
        "retail_price_4",
        "size_archive",
        "created_by",
        "created_at",
        "modified_at",
        "modified_by",
        "barcode_mapping_code",
        "barcode_print_count",
        "barcode_print_time",
        "national_standard_code",
        "national_standard_sequence",
        "is_stopped",
        "is_stop_order",
        "special_price_product",
        "exchangeable_product",
        "exchange_discount_within_period",
        "exchange_discount_outside_period",
        "detail_price_difference",
        "specification_allocation",
        "online_product",
        "barcode_print_info_1",
        "barcode_print_info_2",
        "image_path",
        "small_category_code",
        "small_category_name",
        "supplier_name_2",
        "supplier_name_3",
        "is_sample_clothing",
        "is_sample_clothing_name",
        "main_material_2",
        "main_material_2_name",
        "accessory_1",
        "accessory_1_name",
        "accessory_2",
        "accessory_2_name",
        "filler_1",
        "filler_1_name",
        "safety_technology_category_code",
        "safety_technology_category_name",
        "implementation_standard_code",
        "implementation_standard_name",
        "lining_material",
        "lining_material_name",
        "barcode_print_info_3",
        "barcode_print_info_4",
        "barcode_print_info_5",
        "enabled_batch",
    ],
    "dim_product_option": ["product_code", "option_type", "option_value", "option_name", "source_order"],
    "dim_store": [
        "store_code",
        "store_name",
        "store_type_code",
        "store_type_name",
        "channel_code",
        "region_code",
        "region_name",
        "province",
        "city",
        "open_date",
        "close_date",
        "imported_at",
        "updated_at",
    ],
    "dim_channel": [
        "channel_code",
        "channel_name",
        "channel_category_code",
        "channel_category_name",
        "parent_channel_code",
        "parent_channel_name",
        "region_code",
        "region_name",
        "imported_at",
        "updated_at",
    ],
    "dim_calendar": [
        "date_key",
        "year",
        "month",
        "day",
        "quarter",
        "month_name",
        "day_of_week",
        "day_name",
        "week_of_year",
        "is_weekend",
        "is_month_start",
        "is_month_end",
        "is_quarter_start",
        "is_quarter_end",
        "is_year_start",
        "is_year_end",
    ],
}

TABLE_CONFLICT_TARGETS = {
    "dim_product": ["product_code"],
    "dim_product_option": ["product_code", "option_type", "option_value"],
    "dim_store": ["store_code"],
    "dim_channel": ["channel_code"],
    "dim_calendar": ["date_key"],
}

SALES_HEADER_ROW = 14
SALES_REQUIRED_COLUMNS = [
    "商品代码",
    "颜色代码",
    "颜色名称",
    "尺码代码",
    "尺码名称",
    "商店代码",
    "日期",
    "单据编号",
    "单据类型",
    "数量",
    "选定金额",
    "金额",
]
FACT_RETAIL_SALES_COLUMNS = [
    "id",
    "sale_date",
    "date_key",
    "product_code",
    "color_code",
    "color_name",
    "size_code",
    "size_name",
    "store_code",
    "document_no",
    "document_type",
    "qty",
    "standard_amount",
    "amount",
    "standard_price",
    "unit_price",
    "discount_rate",
    "source_file",
    "source_row_hash",
    "load_batch_id",
    "import_run_time",
    "imported_at",
]


@dataclass(frozen=True)
class ImportResult:
    table_name: str
    rows_read: int
    rows_upserted: int


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _float(value):
    text = _text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _date(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text or None


def _split_detail(value: str) -> list[str]:
    if not value:
        return []
    normalized = value
    for separator in (",", "，", ";", "；", "|", "、", "/"):
        normalized = normalized.replace(separator, ",")
    return [item.strip() for item in normalized.split(",") if item.strip()]


def _safe_hash(parts: list[str]) -> str:
    import hashlib

    digest = hashlib.sha256()
    for part in parts:
        digest.update((part or "").encode("utf-8"))
        digest.update(b"\x1f")
    return digest.hexdigest()


def _load_sheet(path: str | Path, sheet_name: str, header_row: int) -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Workbook {path} is missing sheet {sheet_name}")
    ws = workbook[sheet_name]
    headers: list[str] = []
    records: list[dict[str, str]] = []
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [_text(value) for value in row]
        if row_index == header_row:
            headers = [value for value in values if value]
            continue
        if row_index <= header_row:
            continue
        if not any(values):
            continue
        record = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
        records.append(record)
    return headers, records


def _upsert(conn, table_name: str, columns: list[str], rows: list[tuple], conflict_columns: list[str]) -> int:
    if not rows:
        return 0
    update_columns = [column for column in columns if column not in conflict_columns]
    placeholders = ", ".join(["?"] * len(columns))
    column_sql = ", ".join(columns)
    for row in rows:
        if len(row) != len(columns):
            raise ValueError(
                f"{table_name}: expected {len(columns)} columns but got {len(row)} values. columns={columns}"
            )
    if update_columns:
        update_sql = ", ".join([f"{column}=excluded.{column}" for column in update_columns])
        sql = f"INSERT INTO {table_name} ({column_sql}) VALUES ({placeholders}) ON CONFLICT({', '.join(conflict_columns)}) DO UPDATE SET {update_sql}"
    else:
        sql = f"INSERT INTO {table_name} ({column_sql}) VALUES ({placeholders}) ON CONFLICT({', '.join(conflict_columns)}) DO NOTHING"
    conn.executemany(sql, rows)
    return len(rows)


def _existing_columns(conn, table_name: str) -> list[str]:
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return [row[1] for row in rows]


def _recreate_table_if_needed(conn, table_name: str, create_sql: str) -> None:
    desired_columns = DESIRED_TABLE_COLUMNS[table_name]
    existing = _existing_columns(conn, table_name)
    if existing and existing != desired_columns:
        conn.execute(f"DROP TABLE IF EXISTS {table_name}")
    conn.execute(create_sql)


def create_master_tables(conn) -> None:
    _recreate_table_if_needed(
        conn,
        "dim_product",
        """
        CREATE TABLE IF NOT EXISTS dim_product (
            product_code TEXT PRIMARY KEY,
            product_name TEXT,
            color_detail TEXT,
            size_detail TEXT,
            unit_name TEXT,
            year_code TEXT,
            major_category_code TEXT,
            major_category_name TEXT,
            season_code TEXT,
            season_name TEXT,
            brand_code TEXT,
            brand_name TEXT,
            supplier_code TEXT,
            supplier_name TEXT,
            category_code TEXT,
            category_name TEXT,
            production_factory_code TEXT,
            production_factory_name TEXT,
            designer_code TEXT,
            designer_name TEXT,
            launch_wave_code TEXT,
            launch_wave_name TEXT,
            main_material_code TEXT,
            main_material_name TEXT,
            series_code TEXT,
            series_name TEXT,
            standard_purchase_price REAL,
            purchase_price_1 REAL,
            purchase_price_2 REAL,
            cost_price REAL,
            standard_retail_price REAL,
            foreign_trade_price REAL,
            shipping_price REAL,
            retail_price_3 REAL,
            retail_price_4 REAL,
            size_archive TEXT,
            created_by TEXT,
            created_at TEXT,
            modified_at TEXT,
            modified_by TEXT,
            barcode_mapping_code TEXT,
            barcode_print_count TEXT,
            barcode_print_time TEXT,
            national_standard_code TEXT,
            national_standard_sequence TEXT,
            is_stopped TEXT,
            is_stop_order TEXT,
            special_price_product TEXT,
            exchangeable_product TEXT,
            exchange_discount_within_period TEXT,
            exchange_discount_outside_period TEXT,
            detail_price_difference TEXT,
            specification_allocation TEXT,
            online_product TEXT,
            barcode_print_info_1 TEXT,
            barcode_print_info_2 TEXT,
            image_path TEXT,
            small_category_code TEXT,
            small_category_name TEXT,
            supplier_name_2 TEXT,
            supplier_name_3 TEXT,
            is_sample_clothing TEXT,
            is_sample_clothing_name TEXT,
            main_material_2 TEXT,
            main_material_2_name TEXT,
            accessory_1 TEXT,
            accessory_1_name TEXT,
            accessory_2 TEXT,
            accessory_2_name TEXT,
            filler_1 TEXT,
            filler_1_name TEXT,
            safety_technology_category_code TEXT,
            safety_technology_category_name TEXT,
            implementation_standard_code TEXT,
            implementation_standard_name TEXT,
            lining_material TEXT,
            lining_material_name TEXT,
            barcode_print_info_3 TEXT,
            barcode_print_info_4 TEXT,
            barcode_print_info_5 TEXT,
            enabled_batch TEXT
        );
        """
    )
    _recreate_table_if_needed(
        conn,
        "dim_product_option",
        """
        CREATE TABLE IF NOT EXISTS dim_product_option (
            product_code TEXT NOT NULL,
            option_type TEXT NOT NULL,
            option_value TEXT NOT NULL,
            option_name TEXT,
            source_order INTEGER NOT NULL,
            PRIMARY KEY (product_code, option_type, option_value)
        );
        """,
    )
    _recreate_table_if_needed(
        conn,
        "dim_store",
        """
        CREATE TABLE IF NOT EXISTS dim_store (
            store_code TEXT PRIMARY KEY,
            store_name TEXT,
            store_type_code TEXT,
            store_type_name TEXT,
            channel_code TEXT,
            region_code TEXT,
            region_name TEXT,
            province TEXT,
            city TEXT,
            open_date TEXT,
            close_date TEXT,
            imported_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        """,
    )
    _recreate_table_if_needed(
        conn,
        "dim_channel",
        """
        CREATE TABLE IF NOT EXISTS dim_channel (
            channel_code TEXT PRIMARY KEY,
            channel_name TEXT,
            channel_category_code TEXT,
            channel_category_name TEXT,
            parent_channel_code TEXT,
            parent_channel_name TEXT,
            region_code TEXT,
            region_name TEXT,
            imported_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        """,
    )
    _recreate_table_if_needed(
        conn,
        "dim_calendar",
        """
        CREATE TABLE IF NOT EXISTS dim_calendar (
            date_key TEXT PRIMARY KEY,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            day INTEGER NOT NULL,
            quarter INTEGER NOT NULL,
            month_name TEXT NOT NULL,
            day_of_week INTEGER NOT NULL,
            day_name TEXT NOT NULL,
            week_of_year INTEGER NOT NULL,
            is_weekend INTEGER NOT NULL,
            is_month_start INTEGER NOT NULL,
            is_month_end INTEGER NOT NULL,
            is_quarter_start INTEGER NOT NULL,
            is_quarter_end INTEGER NOT NULL,
            is_year_start INTEGER NOT NULL,
            is_year_end INTEGER NOT NULL
        );
        """,
    )


def import_products(conn, path: str | Path) -> ImportResult:
    _, records = _load_sheet(path, PRODUCT_SHEET, PRODUCT_HEADER_ROW)
    rows = []
    option_rows = []
    for record in records:
        product_code = _text(record.get("商品代码"))
        if not product_code:
            continue
        color_detail = _text(record.get("颜色明细"))
        size_detail = _text(record.get("尺码明细"))
        rows.append(
            (
                product_code,
                _text(record.get("商品名称")),
                color_detail,
                size_detail,
                _text(record.get("单位名称")),
                _text(record.get("年份")),
                _text(record.get("大类")),
                _text(record.get("大类名称")),
                _text(record.get("季节")),
                _text(record.get("季节名称")),
                _text(record.get("品牌")),
                _text(record.get("品牌名称")),
                _text(record.get("供应商")),
                _text(record.get("供应商名称")),
                _text(record.get("类别")),
                _text(record.get("类别名称")),
                _text(record.get("生产工厂")),
                _text(record.get("生产工厂名称")),
                _text(record.get("设计师")),
                _text(record.get("设计师名称")),
                _text(record.get("上货波段")),
                _text(record.get("上货波段名称")),
                _text(record.get("主面料")),
                _text(record.get("主面料名称")),
                _text(record.get("系列")),
                _text(record.get("系列名称")),
                _float(record.get("标准进价")),
                _float(record.get("进价1")),
                _float(record.get("进价2")),
                _float(record.get("成本价")),
                _float(record.get("标准售价")),
                _float(record.get("外贸售价")),
                _float(record.get("发货价")),
                _float(record.get("售价3")),
                _float(record.get("售价4")),
                _text(record.get("尺码档")),
                _text(record.get("建档人")),
                _text(record.get("建档日期")),
                _text(record.get("修改日期")),
                _text(record.get("修改人")),
                _text(record.get("条码对照码")),
                _text(record.get("条码打印次数")),
                _text(record.get("条码打印时间")),
                _text(record.get("国标码")),
                _text(record.get("国标码序号")),
                _text(record.get("停止使用")),
                _text(record.get("停止订货")),
                _text(record.get("特价商品")),
                _text(record.get("可换货商品")),
                _text(record.get("换货期内换货折")),
                _text(record.get("换货期外换货折")),
                _text(record.get("明细异价")),
                _text(record.get("规格分配")),
                _text(record.get("线上商品")),
                _text(record.get("条码打印信息一")),
                _text(record.get("条码打印信息二")),
                _text(record.get("图片")),
                _text(record.get("小类")),
                _text(record.get("小类名称")),
                _text(record.get("供应商名称")),
                _text(record.get("供应商名称")),
                _text(record.get("是否样衣")),
                _text(record.get("是否样衣名称")),
                _text(record.get("主料")),
                _text(record.get("主料名称")),
                _text(record.get("配料1")),
                _text(record.get("配料1名称")),
                _text(record.get("配料2")),
                _text(record.get("配料2名称")),
                _text(record.get("填充物")),
                _text(record.get("填充物名称")),
                _text(record.get("安全技术类别")),
                _text(record.get("安全技术类别名称")),
                _text(record.get("执行标准")),
                _text(record.get("执行标准名称")),
                _text(record.get("里料")),
                _text(record.get("里料名称")),
                _text(record.get("条码打印信息一")),
                _text(record.get("条码打印信息二")),
                _text(record.get("条码打印信息三")),
                _text(record.get("启用批次")),
            )
        )
        colors = _split_detail(color_detail)
        sizes = _split_detail(size_detail)
        for option_order, color in enumerate(colors, start=1):
            option_rows.append((product_code, "color", color, color, option_order))
        for option_order, size in enumerate(sizes, start=1):
            option_rows.append((product_code, "size", size, size, option_order))

    product_columns = [
        "product_code",
        "product_name",
        "color_detail",
        "size_detail",
        "unit_name",
        "year_code",
        "major_category_code",
        "major_category_name",
        "season_code",
        "season_name",
        "brand_code",
        "brand_name",
        "supplier_code",
        "supplier_name",
        "category_code",
        "category_name",
        "production_factory_code",
        "production_factory_name",
        "designer_code",
        "designer_name",
        "launch_wave_code",
        "launch_wave_name",
        "main_material_code",
        "main_material_name",
        "series_code",
        "series_name",
        "standard_purchase_price",
        "purchase_price_1",
        "purchase_price_2",
        "cost_price",
        "standard_retail_price",
        "foreign_trade_price",
        "shipping_price",
        "retail_price_3",
        "retail_price_4",
        "size_archive",
        "created_by",
        "created_at",
        "modified_at",
        "modified_by",
        "barcode_mapping_code",
        "barcode_print_count",
        "barcode_print_time",
        "national_standard_code",
        "national_standard_sequence",
        "is_stopped",
        "is_stop_order",
        "special_price_product",
        "exchangeable_product",
        "exchange_discount_within_period",
        "exchange_discount_outside_period",
        "detail_price_difference",
        "specification_allocation",
        "online_product",
        "barcode_print_info_1",
        "barcode_print_info_2",
        "image_path",
        "small_category_code",
        "small_category_name",
        "supplier_name_2",
        "supplier_name_3",
        "is_sample_clothing",
        "is_sample_clothing_name",
        "main_material_2",
        "main_material_2_name",
        "accessory_1",
        "accessory_1_name",
        "accessory_2",
        "accessory_2_name",
        "filler_1",
        "filler_1_name",
        "safety_technology_category_code",
        "safety_technology_category_name",
        "implementation_standard_code",
        "implementation_standard_name",
        "lining_material",
        "lining_material_name",
        "barcode_print_info_3",
        "barcode_print_info_4",
        "barcode_print_info_5",
        "enabled_batch",
    ]
    option_columns = ["product_code", "option_type", "option_value", "option_name", "source_order"]
    product_count = _upsert(conn, "dim_product", product_columns, rows, conflict_columns=TABLE_CONFLICT_TARGETS["dim_product"])
    option_count = _upsert(conn, "dim_product_option", option_columns, option_rows, conflict_columns=["product_code", "option_type", "option_value"])
    return ImportResult("dim_product", len(records), product_count + option_count)


def import_stores(conn, path: str | Path) -> ImportResult:
    _, records = _load_sheet(path, MASTER_SHEET, MASTER_HEADER_ROW)
    rows = []
    imported_at = datetime.now().isoformat(timespec="seconds")
    for record in records:
        store_code = _text(record.get("代码"))
        if not store_code:
            continue
        rows.append((
            store_code,
            _text(record.get("名称")),
            _text(record.get("类别")),
            _text(record.get("类别名称")),
            _text(record.get("所属渠道")),
            _text(record.get("区域")),
            _text(record.get("区域名称")),
            _text(record.get("省")),
            _text(record.get("市")),
            None,
            _date(record.get("业务截止日期")),
            imported_at,
            imported_at,
        ))
    columns = [
        "store_code",
        "store_name",
        "store_type_code",
        "store_type_name",
        "channel_code",
        "region_code",
        "region_name",
        "province",
        "city",
        "open_date",
        "close_date",
        "imported_at",
        "updated_at",
    ]
    count = _upsert(conn, "dim_store", columns, rows, conflict_columns=["store_code"])
    return ImportResult("dim_store", len(records), count)


def import_channels(conn, path: str | Path) -> ImportResult:
    _, records = _load_sheet(path, MASTER_SHEET, MASTER_HEADER_ROW)
    rows = []
    imported_at = datetime.now().isoformat(timespec="seconds")
    for record in records:
        channel_code = _text(record.get("代码"))
        if not channel_code:
            continue
        rows.append((
            channel_code,
            _text(record.get("名称")),
            _text(record.get("类别")),
            _text(record.get("类别名称")),
            _text(record.get("所属渠道")),
            _text(record.get("所属渠道名称")),
            _text(record.get("区域")),
            _text(record.get("区域名称")),
            imported_at,
            imported_at,
        ))
    columns = [
        "channel_code",
        "channel_name",
        "channel_category_code",
        "channel_category_name",
        "parent_channel_code",
        "parent_channel_name",
        "region_code",
        "region_name",
        "imported_at",
        "updated_at",
    ]
    count = _upsert(conn, "dim_channel", columns, rows, conflict_columns=["channel_code"])
    return ImportResult("dim_channel", len(records), count)


def create_calendar(conn, start_date: str = "2020-01-01", end_date: str = "2035-12-31") -> int:
    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = datetime.strptime(end_date, "%Y-%m-%d").date()
    rows = []
    current = start
    while current <= end:
        month_end = (current.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        rows.append((
            current.isoformat(),
            current.year,
            current.month,
            current.day,
            ((current.month - 1) // 3) + 1,
            current.strftime("%B"),
            current.isoweekday(),
            current.strftime("%A"),
            int(current.strftime("%W")) + 1,
            1 if current.weekday() >= 5 else 0,
            1 if current.day == 1 else 0,
            1 if current == month_end else 0,
            1 if current.month in (1, 4, 7, 10) and current.day == 1 else 0,
            1 if current.month in (3, 6, 9, 12) and current == month_end else 0,
            1 if current.month == 1 and current.day == 1 else 0,
            1 if current.month == 12 and current == month_end else 0,
        ))
        current += timedelta(days=1)
    conn.execute("DELETE FROM dim_calendar")
    conn.executemany(
        "INSERT INTO dim_calendar(date_key, year, month, day, quarter, month_name, day_of_week, day_name, week_of_year, is_weekend, is_month_start, is_month_end, is_quarter_start, is_quarter_end, is_year_start, is_year_end) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        rows,
    )
    return len(rows)


def ensure_sales_table(conn) -> None:
    conn.execute("PRAGMA journal_mode=WAL")
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS fact_retail_sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_date TEXT NOT NULL,
            date_key TEXT NOT NULL,
            product_code TEXT NOT NULL,
            store_code TEXT NOT NULL,
            color_name TEXT,
            size_name TEXT,
            qty REAL NOT NULL,
            amount REAL,
            unit_price REAL,
            discount_rate REAL,
            source_file TEXT NOT NULL,
            source_row_hash TEXT NOT NULL,
            load_batch_id TEXT NOT NULL,
            import_run_time TEXT NOT NULL,
            imported_at TEXT NOT NULL,
            UNIQUE (source_row_hash)
        );

        CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_date_key ON fact_retail_sales(date_key);
        CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_product_code ON fact_retail_sales(product_code);
        CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_store_code ON fact_retail_sales(store_code);
        CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_batch_id ON fact_retail_sales(load_batch_id);
        """
    )
    ensure_sales_indexes(conn)
    ensure_import_log_table(conn)


def ensure_sales_indexes(conn) -> None:
    conn.executescript(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_fact_retail_sales_source_row_hash
            ON fact_retail_sales(source_row_hash);
        CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_sale_date
            ON fact_retail_sales(sale_date);
        CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_date_key
            ON fact_retail_sales(date_key);
        CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_product_code
            ON fact_retail_sales(product_code);
        CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_store_code
            ON fact_retail_sales(store_code);
        CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_product_code_sale_date
            ON fact_retail_sales(product_code, sale_date);
        CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_store_code_sale_date
            ON fact_retail_sales(store_code, sale_date);
        CREATE INDEX IF NOT EXISTS idx_dim_product_product_code
            ON dim_product(product_code);
        CREATE INDEX IF NOT EXISTS idx_dim_product_year_code
            ON dim_product(year_code);
        CREATE INDEX IF NOT EXISTS idx_dim_product_season_name
            ON dim_product(season_name);
        CREATE INDEX IF NOT EXISTS idx_dim_product_launch_wave_name
            ON dim_product(launch_wave_name);
        CREATE INDEX IF NOT EXISTS idx_dim_product_category_name
            ON dim_product(category_name);
        CREATE INDEX IF NOT EXISTS idx_dim_product_major_category_name
            ON dim_product(major_category_name);
        CREATE INDEX IF NOT EXISTS idx_dim_product_designer_name
            ON dim_product(designer_name);
        CREATE INDEX IF NOT EXISTS idx_dim_store_store_code
            ON dim_store(store_code);
        CREATE INDEX IF NOT EXISTS idx_dim_store_region_name
            ON dim_store(region_name);
        CREATE INDEX IF NOT EXISTS idx_dim_store_channel_code
            ON dim_store(channel_code);
        """
    )


def ensure_import_log_table(conn) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS import_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            load_batch_id TEXT,
            import_type TEXT NOT NULL,
            source_file TEXT NOT NULL,
            started_at TEXT NOT NULL,
            finished_at TEXT NOT NULL,
            rows_read INTEGER NOT NULL DEFAULT 0,
            rows_imported INTEGER NOT NULL DEFAULT 0,
            duplicate_rows INTEGER NOT NULL DEFAULT 0,
            unknown_product_rows INTEGER NOT NULL DEFAULT 0,
            unknown_product_codes TEXT,
            unknown_store_rows INTEGER NOT NULL DEFAULT 0,
            unknown_store_codes TEXT,
            elapsed_seconds REAL NOT NULL DEFAULT 0,
            status TEXT NOT NULL,
            message TEXT
        );
        """
    )


def write_import_log(conn, entry: dict[str, object]) -> int:
    columns = [
        "load_batch_id",
        "import_type",
        "source_file",
        "started_at",
        "finished_at",
        "rows_read",
        "rows_imported",
        "duplicate_rows",
        "unknown_product_rows",
        "unknown_product_codes",
        "unknown_store_rows",
        "unknown_store_codes",
        "elapsed_seconds",
        "status",
        "message",
    ]
    values = tuple(entry.get(column) for column in columns)
    conn.execute(
        """
        INSERT INTO import_log (
            load_batch_id, import_type, source_file, started_at, finished_at,
            rows_read, rows_imported, duplicate_rows, unknown_product_rows,
            unknown_product_codes, unknown_store_rows, unknown_store_codes,
            elapsed_seconds, status, message
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        values,
    )
    return int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])


def _find_sales_header_row(ws) -> int:
    normalized_required = {"日期", "商品代码", "数量"}
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [_text(value) for value in row]
        if not any(values):
            continue
        normalized = {value for value in values if value}
        if normalized_required.issubset(normalized):
            return row_index
    raise ValueError("Unable to locate sales header row")


def _load_sales_rows(path: str | Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    header_row = _find_sales_header_row(ws)
    headers: list[str] = []
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [_text(value) for value in row]
        if row_index == header_row:
            headers = [value for value in values if value]
            break
    if not headers:
        raise ValueError(f"Unable to read sales header row from {path}")
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_index <= header_row:
            continue
        values = [_text(value) for value in row]
        if not any(values):
            continue
        record = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
        yield header_row, row_index, record


def _store_lookups(conn) -> tuple[dict[str, str], dict[str, str]]:
    rows = conn.execute("SELECT store_code, store_name FROM dim_store").fetchall()
    name_to_code: dict[str, str] = {}
    code_to_name: dict[str, str] = {}
    for row in rows:
        store_code = _text(row[0])
        store_name = _text(row[1])
        if store_name and store_code and store_name not in name_to_code:
            name_to_code[store_name] = store_code
        if store_code and store_name and store_code not in code_to_name:
            code_to_name[store_code] = store_name
    return name_to_code, code_to_name


def _next_batch_id(conn, run_date: str) -> str:
    prefix = run_date.replace("-", "")
    row = conn.execute(
        "SELECT load_batch_id FROM fact_retail_sales WHERE load_batch_id LIKE ? ORDER BY load_batch_id DESC LIMIT 1",
        (f"{prefix}_%",),
    ).fetchone()
    if not row:
        return f"{prefix}_001"
    suffix = int(str(row[0]).split("_")[-1]) + 1
    return f"{prefix}_{suffix:03d}"


def rebuild_sales_table(conn) -> None:
    conn.execute("DROP TABLE IF EXISTS fact_retail_sales")
    ensure_sales_table(conn)


def import_sales_file(conn, path: str | Path, batch_size: int = 10000) -> dict[str, int | float | str]:
    ensure_sales_table(conn)
    source_path = str(Path(path).resolve())
    import_run_time = datetime.now().isoformat(timespec="seconds")
    run_date = datetime.now().strftime("%Y-%m-%d")
    load_batch_id = _next_batch_id(conn, run_date)
    imported_at = import_run_time

    rows_read = 0
    rows_imported = 0
    duplicate_rows = 0
    unknown_product_codes: set[str] = set()
    unknown_store_codes: set[str] = set()
    unknown_product_rows = 0
    unknown_store_rows = 0
    batch_rows: list[tuple] = []

    product_codes = {row[0] for row in conn.execute("SELECT product_code FROM dim_product").fetchall()}
    store_codes = {row[0] for row in conn.execute("SELECT store_code FROM dim_store").fetchall()}
    store_name_to_code, store_code_to_name = _store_lookups(conn)

    for header_row, excel_row, record in _load_sales_rows(path):
        rows_read += 1
        sale_date = _date(record.get("日期"))
        product_code = _text(record.get("商品代码"))
        color_code = _text(record.get("颜色代码"))
        color_name = _text(record.get("颜色名称"))
        size_code = _text(record.get("尺码代码"))
        size_name = _text(record.get("尺码名称"))
        store_code = _text(record.get("商店代码"))
        document_no = _text(record.get("单据编号"))
        document_type = _text(record.get("单据类型"))
        qty = _float(record.get("数量")) or 0.0
        standard_amount = _float(record.get("选定金额")) or 0.0
        amount = _float(record.get("金额")) or 0.0
        standard_price = standard_amount / qty if qty else 0.0
        unit_price = amount / qty if qty else 0.0
        discount_rate = (amount / standard_amount) if standard_amount else 0.0

        if product_code and product_code not in product_codes:
            unknown_product_rows += 1
            unknown_product_codes.add(product_code)
        if store_code and store_code not in store_codes:
            unknown_store_rows += 1
            unknown_store_codes.add(store_code)

        if not sale_date or not product_code or not store_code:
            continue

        date_key = sale_date.replace("-", "")
        row_hash = _safe_hash([
            sale_date,
            product_code,
            color_code,
            color_name,
            size_code,
            size_name,
            store_code,
            document_no,
            document_type,
            str(qty),
            str(standard_amount),
            str(amount),
            source_path,
        ])
        batch_rows.append((
            sale_date,
            date_key,
            product_code,
            color_code,
            color_name,
            size_code,
            size_name,
            store_code,
            document_no,
            document_type,
            qty,
            standard_amount,
            amount,
            standard_price,
            unit_price,
            discount_rate,
            source_path,
            row_hash,
            load_batch_id,
            import_run_time,
            imported_at,
        ))

        if len(batch_rows) >= batch_size:
            rows_imported += _flush_sales_rows(conn, batch_rows)
            conn.commit()
            batch_rows.clear()

    if batch_rows:
        rows_imported += _flush_sales_rows(conn, batch_rows)
        conn.commit()

    if unknown_product_codes:
        logger.warning("Unknown products: %s", ", ".join(sorted(unknown_product_codes)))
    if unknown_store_codes:
        logger.warning("Unknown stores: %s", ", ".join(sorted(unknown_store_codes)))

    return {
        "source_file": source_path,
        "rows_read": rows_read,
        "rows_imported": rows_imported,
        "duplicate_rows": rows_read - rows_imported,
        "unknown_product_rows": unknown_product_rows,
        "unknown_product_codes": ",".join(sorted(unknown_product_codes)),
        "unknown_store_rows": unknown_store_rows,
        "unknown_store_codes": ",".join(sorted(unknown_store_codes)),
        "unknown_products": len(unknown_product_codes),
        "unknown_stores": len(unknown_store_codes),
        "load_batch_id": load_batch_id,
        "import_run_time": import_run_time,
    }


def _flush_sales_rows(conn, rows: list[tuple]) -> int:
    before = conn.total_changes
    conn.executemany(
        """
        INSERT INTO fact_retail_sales(
            sale_date, date_key, product_code, color_code, color_name, size_code, size_name,
            store_code, document_no, document_type, qty, standard_amount, amount,
            standard_price, unit_price, discount_rate, source_file, source_row_hash,
            load_batch_id, import_run_time, imported_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(source_row_hash) DO UPDATE SET
            sale_date=excluded.sale_date,
            date_key=excluded.date_key,
            product_code=excluded.product_code,
            color_code=excluded.color_code,
            color_name=excluded.color_name,
            size_code=excluded.size_code,
            size_name=excluded.size_name,
            store_code=excluded.store_code,
            document_no=excluded.document_no,
            document_type=excluded.document_type,
            qty=excluded.qty,
            standard_amount=excluded.standard_amount,
            amount=excluded.amount,
            standard_price=excluded.standard_price,
            unit_price=excluded.unit_price,
            discount_rate=excluded.discount_rate,
            source_file=excluded.source_file,
            load_batch_id=excluded.load_batch_id,
            import_run_time=excluded.import_run_time,
            imported_at=excluded.imported_at
        """,
        rows,
    )
    inserted = conn.total_changes - before
    return inserted


def ensure_sales_table(conn) -> None:
    desired_columns = FACT_RETAIL_SALES_COLUMNS
    existing_columns = _existing_columns(conn, "fact_retail_sales")
    if existing_columns and existing_columns != desired_columns:
        conn.execute("DROP TABLE IF EXISTS fact_retail_sales")
        existing_columns = []
    if not existing_columns:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS fact_retail_sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_date TEXT NOT NULL,
                date_key TEXT NOT NULL,
                product_code TEXT NOT NULL,
                color_code TEXT NOT NULL,
                color_name TEXT,
                size_code TEXT NOT NULL,
                size_name TEXT,
                store_code TEXT NOT NULL,
                document_no TEXT NOT NULL,
                document_type TEXT NOT NULL,
                qty REAL NOT NULL,
                standard_amount REAL NOT NULL,
                amount REAL NOT NULL,
                standard_price REAL NOT NULL,
                unit_price REAL NOT NULL,
                discount_rate REAL,
                source_file TEXT NOT NULL,
                source_row_hash TEXT NOT NULL,
                load_batch_id TEXT NOT NULL,
                import_run_time TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                UNIQUE (source_row_hash)
            );
            CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_sale_date ON fact_retail_sales(sale_date);
            CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_date_key ON fact_retail_sales(date_key);
            CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_product_code ON fact_retail_sales(product_code);
            CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_store_code ON fact_retail_sales(store_code);
            CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_product_code_sale_date ON fact_retail_sales(product_code, sale_date);
            CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_store_code_sale_date ON fact_retail_sales(store_code, sale_date);
            CREATE UNIQUE INDEX IF NOT EXISTS idx_fact_retail_sales_source_row_hash ON fact_retail_sales(source_row_hash);
            CREATE INDEX IF NOT EXISTS idx_fact_retail_sales_batch_id ON fact_retail_sales(load_batch_id);
            """
        )
    ensure_import_log_table(conn)


def table_counts(conn) -> dict[str, int]:
    counts = {}
    for table_name in ("dim_product", "dim_product_option", "dim_store", "dim_channel", "dim_calendar"):
        row = conn.execute(f"SELECT COUNT(*) AS count FROM {table_name}").fetchone()
        counts[table_name] = int(row[0]) if row else 0
    return counts


def import_master_data(products: str | None = None, stores: str | None = None, channels: str | None = None, calendar_only: bool = False, calendar_start: str = "2020-01-01", calendar_end: str = "2035-12-31") -> dict[str, int]:
    with get_db_connection() as conn:
        create_master_tables(conn)
        results: dict[str, int] = {}
        if not calendar_only:
            if products:
                result = import_products(conn, products)
                logger.info("Imported %s rows into %s", result.rows_upserted, result.table_name)
                results[result.table_name] = result.rows_upserted
            if stores:
                result = import_stores(conn, stores)
                logger.info("Imported %s rows into %s", result.rows_upserted, result.table_name)
                results[result.table_name] = result.rows_upserted
            if channels:
                result = import_channels(conn, channels)
                logger.info("Imported %s rows into %s", result.rows_upserted, result.table_name)
                results[result.table_name] = result.rows_upserted
        calendar_rows = create_calendar(conn, calendar_start, calendar_end)
        logger.info("Imported %s rows into dim_calendar", calendar_rows)
        results["dim_calendar"] = calendar_rows
        conn.commit()
        return results
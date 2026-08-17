from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from database import get_db_connection
from importers.master_data_importer import _existing_columns, _float, _text, ensure_import_log_table
from logging_config import get_logger


logger = get_logger(__name__)

INVENTORY_REQUIRED_COLUMNS = ["商品代码", "仓库代码", "颜色名称", "尺码名称", "数量"]
INVENTORY_DATE_PATTERN = "日期:"
INVENTORY_TABLE = "fact_inventory_snapshot"


@dataclass(frozen=True)
class InventoryImportResult:
    source_file: str
    inventory_date: str
    rows_read: int
    rows_imported: int
    unique_products: int
    unique_warehouses: int
    positive_inventory_rows: int
    negative_inventory_rows: int
    net_inventory_quantity: float
    available_inventory_quantity: float
    unmatched_warehouse_count: int


def _find_header_row(ws) -> int:
    required = set(INVENTORY_REQUIRED_COLUMNS)
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [_text(value) for value in row]
        if not any(values):
            continue
        if required.issubset({value for value in values if value}):
            return row_index
    raise ValueError("Unable to locate inventory header row")


def _extract_inventory_date_from_filename(path: str | Path) -> str:
    match = re.search(r"(20\d{6})", Path(path).name)
    if not match:
        raise ValueError("Inventory date is missing")
    return datetime.strptime(match.group(1), "%Y%m%d").date().isoformat()


def _extract_inventory_date(ws, path: str | Path | None = None) -> str:
    for row_index in range(1, 9):
        for cell in ws[row_index]:
            value = _text(cell.value)
            if not value:
                continue
            if value.startswith(INVENTORY_DATE_PATTERN):
                inventory_date = value.split(INVENTORY_DATE_PATTERN, 1)[1].strip()
                if inventory_date:
                    return inventory_date
    if path is not None:
        return _extract_inventory_date_from_filename(path)
    raise ValueError("Inventory date is missing")


def _ensure_inventory_table(conn) -> None:
    desired_columns = [
        "inventory_date",
        "product_code",
        "warehouse_code",
        "color_name",
        "size_name",
        "raw_inventory_qty",
        "available_inventory_qty",
        "source_file",
        "imported_at",
    ]
    existing = _existing_columns(conn, INVENTORY_TABLE)
    if existing and existing != desired_columns:
        conn.execute(f"DROP TABLE IF EXISTS {INVENTORY_TABLE}")
        existing = []
    if not existing:
        conn.executescript(
            f"""
            CREATE TABLE IF NOT EXISTS {INVENTORY_TABLE} (
                inventory_date TEXT NOT NULL,
                product_code TEXT NOT NULL,
                warehouse_code TEXT NOT NULL,
                color_name TEXT NOT NULL DEFAULT '',
                size_name TEXT NOT NULL DEFAULT '',
                raw_inventory_qty REAL NOT NULL DEFAULT 0,
                available_inventory_qty REAL NOT NULL DEFAULT 0,
                source_file TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                UNIQUE (inventory_date, product_code, warehouse_code, color_name, size_name)
            );
            CREATE INDEX IF NOT EXISTS idx_fact_inventory_snapshot_date ON {INVENTORY_TABLE}(inventory_date);
            CREATE INDEX IF NOT EXISTS idx_fact_inventory_snapshot_product ON {INVENTORY_TABLE}(product_code);
            CREATE INDEX IF NOT EXISTS idx_fact_inventory_snapshot_warehouse ON {INVENTORY_TABLE}(warehouse_code);
            """
        )


def _flush_rows(conn, rows: list[tuple[Any, ...]]) -> int:
    if not rows:
        return 0
    before = conn.total_changes
    conn.executemany(
        f"""
        INSERT INTO {INVENTORY_TABLE} (
            inventory_date, product_code, warehouse_code, color_name, size_name,
            raw_inventory_qty, available_inventory_qty, source_file, imported_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(inventory_date, product_code, warehouse_code, color_name, size_name) DO UPDATE SET
            raw_inventory_qty=excluded.raw_inventory_qty,
            available_inventory_qty=excluded.available_inventory_qty,
            source_file=excluded.source_file,
            imported_at=excluded.imported_at
        """,
        rows,
    )
    return conn.total_changes - before


def _read_csv_rows(path: str | Path) -> list[list[str]]:
    encodings = ("utf-8-sig", "gb18030", "gbk")
    last_error: UnicodeDecodeError | None = None
    for encoding in encodings:
        try:
            with Path(path).open("r", encoding=encoding, newline="") as handle:
                return [[_text(cell) for cell in row] for row in csv.reader(handle)]
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"Unable to decode inventory CSV file {path}") from last_error


def _merged_header_cells(first_row: list[str], second_row: list[str] | None = None) -> list[str]:
    if second_row is None:
        return first_row
    width = max(len(first_row), len(second_row))
    merged: list[str] = []
    for index in range(width):
        first_value = first_row[index] if index < len(first_row) else ""
        second_value = second_row[index] if index < len(second_row) else ""
        merged.append(first_value or second_value)
    return merged


def _load_inventory_rows_csv(path: str | Path) -> tuple[str, list[str], list[list[str]], int]:
    rows = _read_csv_rows(path)
    required = set(INVENTORY_REQUIRED_COLUMNS)
    header_row = 0
    headers: list[str] = []
    data_start_index = 0
    for index, values in enumerate(rows):
        if not any(values):
            continue
        candidates = [(values, index + 1, index + 1)]
        if index + 1 < len(rows):
            candidates.append((_merged_header_cells(values, rows[index + 1]), index + 1, index + 2))
        for candidate, candidate_header_row, candidate_data_start in candidates:
            normalized = {value for value in candidate if value}
            if required.issubset(normalized):
                header_row = candidate_header_row
                headers = candidate
                data_start_index = candidate_data_start
                break
        if header_row:
            break
    if not header_row:
        raise ValueError("Unable to locate inventory header row")
    missing = sorted(required - {value for value in headers if value})
    if missing:
        raise ValueError(f"Inventory CSV is missing required columns: {', '.join(missing)}")
    return _extract_inventory_date_from_filename(path), headers, rows[data_start_index:], header_row


def _load_inventory_rows_workbook(path: str | Path) -> tuple[str, list[str], Any, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    header_row = _find_header_row(ws)
    headers = [_text(cell.value) for cell in ws[header_row]]
    required_set = set(INVENTORY_REQUIRED_COLUMNS)
    if not required_set.issubset({value for value in headers if value}):
        missing = sorted(required_set - {value for value in headers if value})
        raise ValueError(f"Inventory workbook is missing required columns: {', '.join(missing)}")
    return _extract_inventory_date(ws, path), headers, ws.iter_rows(min_row=header_row + 1, values_only=True), header_row


def import_inventory_file(path: str | Path, batch_size: int = 10000) -> dict[str, Any]:
    source_file = str(Path(path).resolve())
    if Path(path).suffix.lower() == ".csv":
        inventory_date, headers, source_rows, _ = _load_inventory_rows_csv(path)
    else:
        inventory_date, headers, source_rows, _ = _load_inventory_rows_workbook(path)
    imported_at = datetime.now().isoformat(timespec="seconds")

    with get_db_connection() as conn:
        ensure_import_log_table(conn)
        _ensure_inventory_table(conn)

        warehouse_codes = {str(row[0]) for row in conn.execute("SELECT warehouse_code FROM dim_warehouse").fetchall()}
        if not warehouse_codes:
            raise ValueError("dim_warehouse is empty; import warehouses before inventory")

        rows_read = 0
        rows_imported = 0
        positive_inventory_rows = 0
        negative_inventory_rows = 0
        unmatched_warehouse_count = 0
        net_inventory_quantity = 0.0
        available_inventory_quantity = 0.0
        unique_products: set[str] = set()
        unique_warehouses: set[str] = set()
        unmatched_warehouses: set[str] = set()
        batch_rows: list[tuple[Any, ...]] = []

        for values in source_rows:
            values = [_text(value) for value in values]
            if not any(values):
                continue
            record = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
            product_code = _text(record.get("商品代码"))
            warehouse_code = _text(record.get("仓库代码"))
            color_name = _text(record.get("颜色名称"))
            size_name = _text(record.get("尺码名称"))
            raw_qty = _float(record.get("数量")) or 0.0

            rows_read += 1
            if raw_qty > 0:
                positive_inventory_rows += 1
            elif raw_qty < 0:
                negative_inventory_rows += 1
            net_inventory_quantity += raw_qty
            available_inventory_quantity += max(raw_qty, 0.0)
            if product_code:
                unique_products.add(product_code)
            if warehouse_code:
                unique_warehouses.add(warehouse_code)
            if warehouse_code not in warehouse_codes:
                unmatched_warehouses.add(warehouse_code)

            batch_rows.append((
                inventory_date,
                product_code,
                warehouse_code,
                color_name,
                size_name,
                raw_qty,
                max(raw_qty, 0.0),
                source_file,
                imported_at,
            ))

            if len(batch_rows) >= batch_size:
                rows_imported += _flush_rows(conn, batch_rows)
                batch_rows.clear()

        if batch_rows:
            rows_imported += _flush_rows(conn, batch_rows)

        if rows_read == 0:
            raise ValueError("Inventory workbook contains no data rows")
        if unmatched_warehouses:
            raise ValueError(f"Inventory workbook contains unknown warehouse codes: {', '.join(sorted(unmatched_warehouses))}")

        result = {
            "source_file": source_file,
            "inventory_date": inventory_date,
            "rows_read": rows_read,
            "rows_imported": rows_imported,
            "unique_products": len(unique_products),
            "unique_warehouses": len(unique_warehouses),
            "positive_inventory_rows": positive_inventory_rows,
            "negative_inventory_rows": negative_inventory_rows,
            "net_inventory_quantity": net_inventory_quantity,
            "available_inventory_quantity": available_inventory_quantity,
            "unmatched_warehouse_count": len(unmatched_warehouses),
        }
        conn.commit()
        return result
